import marimo

__generated_with = "0.21.1"
app = marimo.App(width="medium")


@app.cell
def _():
    import openpyxl

    # data_only=False ensures we read the formula, not just the calculated value
    workbook = openpyxl.load_workbook(
        "/root/repos/aer/components/aer/data/Satellites 2026-04-01.xlsx",
        data_only=False,
    )
    sheet = workbook.active
    return (sheet,)


@app.cell
def _(sheet):
    # Create a dictionary mapping the column name to its column number (1-indexed)
    column_map = {}
    for cell in sheet[1]:
        if cell.value:
            column_map[cell.value] = cell.column

    print(f"Column Map: {column_map}")
    return


@app.function
def extract_hyperlink_data(sheet, column_name, max_rows=None):
    """
    Extract hyperlink URLs from a specific column in an Excel sheet.

    Why this is needed:
    - Pandas' read_excel() only reads the calculated/display values of cells
    - It does NOT preserve hyperlinks or their underlying URLs
    - openpyxl provides access to the cell.hyperlink attribute which contains
      the actual URL in the .target property

    Args:
        sheet: openpyxl worksheet object
        column_name: Name of the column to extract (must match header in row 1)
        max_rows: Maximum number of rows to process (None for all)

    Returns:
        List of dicts with keys: 'row', 'value', 'url'
        - row: 1-indexed row number
        - value: The display text of the cell
        - url: The hyperlink URL (None if no hyperlink exists)
    """
    # Find the column index from the header
    column_idx = None
    for cell in sheet[1]:
        if cell.value == column_name:
            column_idx = cell.column
            break

    if column_idx is None:
        raise ValueError(f"Column '{column_name}' not found in sheet headers")

    results = []

    # Determine row range
    end_row = max_rows + 1 if max_rows else sheet.max_row + 1
    if end_row > sheet.max_row + 1:
        end_row = sheet.max_row + 1

    # Iterate through rows
    for row_idx in range(2, end_row):
        cell = sheet.cell(row=row_idx, column=column_idx)

        # Get cell value
        value = cell.value

        # Get hyperlink URL if it exists
        url = None
        if cell.hyperlink and cell.hyperlink.target:
            url = cell.hyperlink.target

        results.append({"row": row_idx, "value": value, "url": url})

    return results


@app.cell
def _(sheet):
    # Extract data from Payload column
    print("=" * 60)
    print("PAYLOAD COLUMN EXTRACTION")
    print("=" * 60)

    payload_data = extract_hyperlink_data(sheet, "Payload", max_rows=20)

    # Count hyperlinks
    payload_with_links = [d for d in payload_data if d["url"] is not None]
    payload_without_links = [d for d in payload_data if d["url"] is None]

    print("\nFirst 10 rows:")
    for _item in payload_data[:10]:
        _url_display = (
            _item["url"][:50] + "..."
            if _item["url"] and len(_item["url"]) > 50
            else _item["url"]
        )
        print(f"  Row {_item['row']}: {_item['value']!r}")
        print(f"           URL: {_url_display}")

    print("\nSummary:")
    print(f"  Total rows processed: {len(payload_data)}")
    print(f"  Rows WITH hyperlinks: {len(payload_with_links)}")
    print(f"  Rows WITHOUT hyperlinks: {len(payload_without_links)}")
    return


@app.cell
def _(sheet):
    # Extract data from Acronym column
    print("=" * 60)
    print("ACRONYM COLUMN EXTRACTION")
    print("=" * 60)

    acronym_data = extract_hyperlink_data(sheet, "Acronym", max_rows=20)

    # Count hyperlinks
    acronym_with_links = [d for d in acronym_data if d["url"] is not None]
    acronym_without_links = [d for d in acronym_data if d["url"] is None]

    print("\nFirst 10 rows:")
    for _item in acronym_data[:10]:
        _url_display = (
            _item["url"][:50] + "..."
            if _item["url"] and len(_item["url"]) > 50
            else _item["url"]
        )
        print(f"  Row {_item['row']}: {_item['value']!r}")
        print(f"           URL: {_url_display}")

    print("\nSummary:")
    print(f"  Total rows processed: {len(acronym_data)}")
    print(f"  Rows WITH hyperlinks: {len(acronym_with_links)}")
    print(f"  Rows WITHOUT hyperlinks: {len(acronym_without_links)}")
    return


@app.cell
def _():
    print("=" * 60)
    print("NOTES ON HYPERLINK EXTRACTION")
    print("=" * 60)
    print("""
    Why pandas can't read hyperlinks:
    - pd.read_excel() only extracts the DISPLAY text from cells
    - Excel hyperlinks are stored separately from the cell value
    - The actual URL is in a .rels file within the .xlsx structure
    - openpyxl provides cell.hyperlink to access this data

    Current situation:
    The exported Excel file from OSCAR website doesn't contain hyperlinks.
    This appears to be a limitation of the website's Excel export function.
    The hyperlinks visible on the website table aren't preserved in the export.

    If you need the URLs:
    1. The website links likely follow a pattern like:
       https://space.oscar.wmo.int/satellites/{satellite_id}
       https://space.oscar.wmo.int/instruments/{instrument_id}

    2. You could scrape the data directly from the HTML table on the website
       using requests + BeautifulSoup or similar tools

    3. Or manually construct URLs based on the Acronym/Payload names

    The extraction function above will work correctly when hyperlinks ARE present
    in the Excel file.
    """)
    return


@app.cell
def _():
    import requests
    import pandas as pd
    from typing import List, Dict, Any, Optional
    import time

    BASE_URL = "https://space.oscar.wmo.int/api/v1"

    def fetch_oscar_satellites(max_pages: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Fetch all satellites from OSCAR API with automatic pagination handling.

        API Endpoint: GET /api/v1/satellites
        Documentation: https://space.oscar.wmo.int/apidoc/

        Args:
            max_pages: Maximum number of pages to fetch (None for all)

        Returns:
            List of satellite dictionaries from _embedded.satellites
        """
        satellites = []
        page = 1

        while True:
            url = f"{BASE_URL}/satellites?page={page}"
            print(f"Fetching page {page}...", end=" ")

            try:
                response = requests.get(url, timeout=30)
                response.raise_for_status()
                data = response.json()

                # Extract satellites from _embedded
                page_satellites = data.get("_embedded", {}).get("satellites", [])
                if not page_satellites:
                    print("No more data")
                    break

                satellites.extend(page_satellites)
                print(f"Got {len(page_satellites)} satellites")

                # Check for next page
                next_link = data.get("_links", {}).get("next", {}).get("href")
                if not next_link:
                    print("No next page")
                    break

                page += 1

                # Check max_pages limit
                if max_pages and page > max_pages:
                    print(f"Reached max_pages limit ({max_pages})")
                    break

                # Small delay to be nice to the API
                time.sleep(0.1)

            except requests.exceptions.RequestException as e:
                print(f"Error: {e}")
                break

        return satellites

    def fetch_oscar_instruments(
        max_pages: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        """
        Fetch all instruments from OSCAR API with automatic pagination handling.

        API Endpoint: GET /api/v1/instruments
        Documentation: https://space.oscar.wmo.int/apidoc/

        Args:
            max_pages: Maximum number of pages to fetch (None for all)

        Returns:
            List of instrument dictionaries from _embedded.instruments
        """
        instruments = []
        page = 1

        while True:
            url = f"{BASE_URL}/instruments?page={page}"
            print(f"Fetching instruments page {page}...", end=" ")

            try:
                response = requests.get(url, timeout=30)
                response.raise_for_status()
                data = response.json()

                # Extract instruments from _embedded
                page_instruments = data.get("_embedded", {}).get("instruments", [])
                if not page_instruments:
                    print("No more data")
                    break

                instruments.extend(page_instruments)
                print(f"Got {len(page_instruments)} instruments")

                # Check for next page
                next_link = data.get("_links", {}).get("next", {}).get("href")
                if not next_link:
                    print("No next page")
                    break

                page += 1

                # Check max_pages limit
                if max_pages and page > max_pages:
                    print(f"Reached max_pages limit ({max_pages})")
                    break

                # Small delay to be nice to the API
                time.sleep(0.1)

            except requests.exceptions.RequestException as e:
                print(f"Error: {e}")
                break

        return instruments

    return Dict, List, fetch_oscar_instruments, fetch_oscar_satellites, pd


@app.cell
def _(fetch_oscar_satellites):
    # Fetch satellites from OSCAR API (limited to 2 pages for demo)
    print("=" * 60)
    print("FETCHING SATELLITES FROM OSCAR API")
    print("=" * 60)

    # For demo, fetch only first 2 pages (60 satellites)
    # Remove max_pages parameter to fetch all
    satellites_raw = fetch_oscar_satellites(max_pages=2)

    print(f"\nTotal satellites fetched: {len(satellites_raw)}")
    return (satellites_raw,)


@app.cell
def _(Dict, List, pd, satellites_raw):
    def satellites_to_dataframe(satellites_data: List[Dict]) -> pd.DataFrame:
        """
        Convert raw satellite API data to a clean pandas DataFrame.

        Columns:
            - id: Satellite ID
            - slug: URL-friendly identifier
            - acronym: Short name
            - fullname: Full satellite name
            - space_agency: Operating agency (NASA, JAXA, etc.)
            - status: Active/Inactive/etc.
            - orbit: Orbit type
            - launch_date: Launch date string
            - eol: End of life date
            - altitude: Orbital altitude
            - longitude: Orbital longitude (null for LEO)
            - ect: Equator crossing time
            - wigos_id: WIGOS station identifier
            - data_access_link: URL for data access
            - link: Satellite info link
        """
        if not satellites_data:
            return pd.DataFrame()

        records = []
        for sat in satellites_data:
            record = {
                "id": sat.get("id"),
                "slug": sat.get("slug"),
                "acronym": sat.get("acronym"),
                "fullname": sat.get("fullname"),
                "space_agency": sat.get("space_agency"),
                "status": sat.get("status"),
                "orbit": sat.get("orbit"),
                "launch_date": sat.get("launch_date"),
                "eol": sat.get("EoL"),
                "altitude": sat.get("Altitude"),
                "longitude": sat.get("longitude"),
                "ect": sat.get("ECT"),
                "wigos_id": sat.get("WIGOS_Station_Identifier"),
                "data_access_link": sat.get("data_access_link"),
                "link": sat.get("link"),
            }
            records.append(record)

        return pd.DataFrame(records)

    # Create satellites DataFrame
    df_satellites = satellites_to_dataframe(satellites_raw)

    print("SATELLITES DATAFRAME")
    print("=" * 60)
    print(f"\nShape: {df_satellites.shape}")
    print(f"\nColumns: {list(df_satellites.columns)}")
    print("\nFirst 5 rows:")
    print(df_satellites.head())
    print("\nDataFrame Info:")
    print(df_satellites.info())
    return (df_satellites,)


@app.cell
def _(Dict, List, pd, satellites_raw):
    def extract_instruments_from_satellites(
        satellites_data: List[Dict],
    ) -> pd.DataFrame:
        """
        Extract instruments from satellite data and create a DataFrame.

        Each satellite has a 'satellite-instruments' array containing
        instruments mounted on that satellite.

        Columns:
            - satellite_id: Parent satellite ID (foreign key)
            - satellite_acronym: Parent satellite acronym
            - instrument_id: Instrument ID
            - slug: Instrument slug
            - name: Short instrument name
            - fullname: Full instrument name
            - instrument_type: Type of instrument
            - providing_agency: Agency providing the instrument
            - start_date: When instrument started operating
            - eol: End of life for instrument
            - status: Instrument status
            - classification: List of classification categories (flattened)
        """
        if not satellites_data:
            return pd.DataFrame()

        records = []
        for sat in satellites_data:
            sat_id = sat.get("id")
            sat_acronym = sat.get("acronym")
            instruments = sat.get("satellite-instruments", [])

            for inst in instruments:
                instrument_data = inst.get("instrument", {})

                record = {
                    "satellite_id": sat_id,
                    "satellite_acronym": sat_acronym,
                    "instrument_id": instrument_data.get("id"),
                    "slug": instrument_data.get("slug"),
                    "name": instrument_data.get("name"),
                    "fullname": instrument_data.get("fullname"),
                    "instrument_type": instrument_data.get("instrumenttype"),
                    "providing_agency": instrument_data.get("providing-agency"),
                    "start_date": inst.get("start-date"),
                    "eol": inst.get("EoL"),
                    "status": inst.get("status"),
                    "classification": ", ".join(inst.get("classification", []))
                    if inst.get("classification")
                    else None,
                }
                records.append(record)

        return pd.DataFrame(records)

    # Create instruments DataFrame
    df_instruments = extract_instruments_from_satellites(satellites_raw)

    print("INSTRUMENTS DATAFRAME")
    print("=" * 60)
    print(f"\nShape: {df_instruments.shape}")
    print(f"\nColumns: {list(df_instruments.columns)}")
    print("\nFirst 10 rows:")
    print(df_instruments.head(10))
    print("\nInstruments per satellite:")
    print(df_instruments.groupby("satellite_acronym").size())
    return (df_instruments,)


@app.cell
def _(df_instruments, df_satellites):
    print("=" * 60)
    print("SUMMARY STATISTICS")
    print("=" * 60)

    print("\nSatellites:")
    print(f"  Total: {len(df_satellites)}")
    print(f"  Active: {len(df_satellites[df_satellites['status'] == 'Active'])}")
    print(f"  Inactive: {len(df_satellites[df_satellites['status'] == 'Inactive'])}")
    print(f"  Space Agencies: {df_satellites['space_agency'].nunique()}")
    print("  Top agencies:")
    print(df_satellites["space_agency"].value_counts().head())

    print("\nInstruments:")
    print(f"  Total: {len(df_instruments)}")
    print(f"  Unique instruments: {df_instruments['instrument_id'].nunique()}")
    print("  Instrument types:")
    print(df_instruments["instrument_type"].value_counts().head(10))

    print("\nSample satellite-instrument relationships:")
    sample = df_instruments[["satellite_acronym", "name", "instrument_type"]].head(10)
    print(sample.to_string(index=False))
    return


@app.cell
def _(fetch_oscar_instruments, pd):
    print("=" * 60)
    print("FETCHING ALL INSTRUMENTS FROM OSCAR API")
    print("=" * 60)

    # Fetch all instruments (limited to 2 pages for demo)
    instruments_raw = fetch_oscar_instruments(max_pages=2)

    print(f"\nTotal instruments fetched: {len(instruments_raw)}")

    # Convert to DataFrame
    if instruments_raw:
        df_all_instruments = pd.DataFrame(instruments_raw)
        print(f"\nAll Instruments DataFrame shape: {df_all_instruments.shape}")
        print(f"Columns: {list(df_all_instruments.columns)}")
        print("\nFirst 5 rows:")
        print(df_all_instruments.head())
    else:
        df_all_instruments = pd.DataFrame()
    return


if __name__ == "__main__":
    app.run()
